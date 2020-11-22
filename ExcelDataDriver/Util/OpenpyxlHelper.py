from openpyxl import load_workbook
from openpyxl import Workbook

class OpenpyxlHelper(object):

    @staticmethod
    def load_excel_file(filename, data_only=True, keep_vba=False):
        return load_workbook(filename=filename, data_only=data_only, keep_vba=keep_vba, read_only=False, keep_links=False)

    @staticmethod
    def save_excel_file(filename, source_wb):
        # source_wb['Sheet1'].cell(row=2, column=1).value = 2
        source_wb.save(filename)
